print()
# train.xlsx 읽어서 정규식 적용

from openpyxl import load_workbook
import re
from openpyxl import Workbook

path = "./RPAbasic/crawl/download/"
wb = load_workbook(path + "train.xlsx")
ws = wb.active

# 전체 데이터를 읽어서 성별에 따른 고객명단 이동
# 새로운 엑셀파일(train_gender.xlsx) 작성 - 4개 시트 생성
# 엑셀 파일 생성
work_book = Workbook()

# 첫번째 시트 생성
work_sheet_man = work_book.active
work_sheet_man.title = "남성"
work_sheet_man.column_dimensions["D"].width = 70

# 두번째 시트 (만들어야함)
work_sheet_solo_women = work_book.create_sheet()
work_sheet_solo_women.title = "미혼 여성"
work_sheet_solo_women.column_dimensions["D"].width = 70

# 세번째 시트 (만들어야함)
work_sheet_married_women = work_book.create_sheet()
work_sheet_married_women.title = "기혼 여성"
work_sheet_married_women.column_dimensions["D"].width = 70

# 네번째 시트 (만들어야함)
work_sheet_others = work_book.create_sheet()
work_sheet_others.title = "기타"
# 다른 방법 : work_sheet_others = work_book.create_sheet(title="기타")
work_sheet_others.column_dimensions["D"].width = 70

# 5번째 시트 생성(보고서)
# 분류  생존자수 사망자수 생존률
# 남성     48      12     80%
# 미혼여성 ##
# 기혼여성 ##
# 기타     ##

# 5번째 시트 생성
work_sheet_report = work_book.create_sheet(title="보고서")
work_sheet_report.append(["분류", "생존자수", "사망자수", "생존률"])

pattern = re.compile(" [A-Za-z]+\.")

# 생존자수, 사망자 수 카운트 할 변수 선언
man_survived, man_unservived = 0, 0
solo_survived, solo_unservived = 0, 0
married_survived, married_unservived = 0, 0
others_survived, others_unservived = 0, 0


# 남성(Mr.), 미혼여성(Miss.), 기혼 여성(Mrs.), 기타
pattern = re.compile("[A-Za-z]+\.")

for each_row in ws.iter_rows():
    # print()

    # 첫번째 행인 경우 제목행이기 때문에 모든 sheet에 붙여넣기
    title_list = []
    if each_row[0].row == 1:
        # for col in each_row:
        #     title_list.append(col.value)
        #     print(title_list)

        # work_sheet_man.append(title_list)
        # work_sheet_solo_women.append(title_list)
        # work_sheet_married_women.append(title_list)
        # work_sheet_others.append(title_list)

        # 위의 작업을 한번에 실행. List Comprehension 방법
        work_sheet_man.append(col.value for col in each_row)
        work_sheet_solo_women.append(col.value for col in each_row)
        work_sheet_married_women.append(col.value for col in each_row)
        work_sheet_others.append(col.value for col in each_row)

    # 두번째 ~ 마지막행 까지는 이름 읽어서 각 성별에 맞춰서 시트에 붙여넣기
    else:
        # 패턴을 찾는 경우 찾은 문자열 리턴, 못찾으면 비어있는 리스트 [] 리턴
        data = pattern.findall(each_row[3].value)

        if len(data) > 0:
            if data[0] == " Mr.":  # 앞에 공백
                work_sheet_man.append(col.value for col in each_row)

                if each_row[1].value == 1:  # 생존자
                    man_survived += 1
                else:
                    man_unservived += 1

            elif data[0] == " Miss.":  # 앞에 공백
                work_sheet_solo_women.append(col.value for col in each_row)

                if each_row[1].value == 1:  # 생존자
                    solo_survived += 1
                else:
                    solo_unservived += 1

            elif data[0] == " Mrs.":  # 앞에 공백
                work_sheet_married_women.append(col.value for col in each_row)

                if each_row[1].value == 1:  # 생존자
                    married_survived += 1
                else:
                    married_unservived += 1

            else:
                work_sheet_others.append(col.value for col in each_row)

                if each_row[1].value == 1:  # 생존자
                    others_survived += 1
                else:
                    others_unservived += 1


# 보고서 작성
# 남성 생존율
man_survived_rate = "%.2f%%" % (man_survived / (man_survived + man_unservived) * 100)

# 미혼 여성 생존율
solo_survived_rate = "%.2f%%" % (
    solo_survived / (solo_survived + solo_unservived) * 100
)
# 기혼 여성 생존율
married_survived_rate = "%.2f%%" % (
    married_survived / (married_survived + married_unservived) * 100
)
# 기타 생존율
others_survived_rate = "%.2f%%" % (
    others_survived / (others_survived + others_unservived) * 100
)

work_sheet_report.append(["남성", man_survived, man_unservived, man_survived_rate])
work_sheet_report.append(["미혼 여성", solo_survived, solo_unservived, solo_survived_rate])
work_sheet_report.append(
    ["기혼 여성", married_survived, married_unservived, married_survived_rate]
)
work_sheet_report.append(
    ["기타", others_survived, others_unservived, others_survived_rate]
)


# 새로운 엑셀 파일에 저장
work_book.save(path + "train_gender.xlsx")

# 원본 엑셀 닫기
wb.close()

# 새로운 엑셀 파일 닫기
work_book.close()

print()
